from netmiko import ConnectHandler
import openpyxl
from threading import Thread

def create_dic_device(database):
    List_ath = []
    for row in range(2, database.max_row +1):
        dict_ath ={}
        dict_ath["device_type"] = database.cell(row=row,column=1).value
        dict_ath["ip"] = database.cell(row=row,column=2).value
        dict_ath["username"] = database.cell(row=row,column=3).value
        dict_ath["password"] = database.cell(row=row,column=4).value
        dict_ath["port"] = database.cell(row=row,column=5).value
        if dict_ath["device_type"] == None or dict_ath["ip"] == None or dict_ath["username"] == None:
            continue
        List_ath.append(dict_ath)
    return List_ath

def running_config(ath):
    try:
        print(f'connecting to {ath['ip']}')
        connect_router = ConnectHandler(**ath)
        command = connect_router.send_command('show runn')
        with open(f'{ath['device_type']}-{ath['ip']}:{ath['port']}', 'w') as config_file:
            config_file.write(command)
    except Exception as e:
        print(f'ERROR => {e}')

source_address = openpyxl.load_workbook('cisco_config_backup_source.xlsx')
db = source_address.active
device_list = create_dic_device(db)
threads = []
for device in device_list:
    t = Thread(target=running_config, args=(device,))
    t.start()
    threads.append(t)

for th in threads:
    th.join()
